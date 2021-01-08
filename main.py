# This is a sample Python script.

# Press <no shortcut> to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
import os
import shutil
import sys
import time
from copy import deepcopy
from copy import copy

import openpyxl as openpyxl


def print_hi(name):
    # path = '/Users/earl/Workspace/yuanli/2021.01'
    try:
        path = sys.argv[1]
        files = os.listdir(path)
        stat_file = ''
        stat_workbook = None
        stat_worksheet = None
        stat_worksheet_init_row = 3
        local_time = time.strftime("%Y-%m-%d", time.localtime())
        file_name = path + "/" + "统计汇总表_" + local_time + ".xlsx"
        if os.path.exists(file_name):
            os.remove(file_name)
        for file in files:
            if not os.path.isdir(file):
                print(path + "/" + file)
                if file.startswith("~$") or file.startswith("统计汇总表"):
                    print(path + "/" + file + '  忽略！！！')
                elif os.path.splitext(file)[-1] == '.xls':
                    print(path + "/" + file + '  格式错误！！！')
                elif os.path.splitext(file)[-1] == '.xlsx':
                    if stat_file == '':
                        stat_file = file_name
                        shutil.copyfile(path + "/" + file, stat_file)
                        stat_workbook = openpyxl.load_workbook(stat_file)
                        # 删除所有数据，只保留表头
                        stat_workbook.remove(stat_workbook.worksheets[0])
                        stat_workbook.remove(stat_workbook.worksheets[1])
                        # stat_worksheet = stat_workbook['每日进度']
                        stat_worksheet = stat_workbook.worksheets[0]
                        m_list = stat_worksheet.merged_cells
                        cr = []
                        for m_area in m_list:
                            # 合并单元格的起始行坐标、终止行坐标。。。。，
                            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
                            # 纵向合并单元格的位置信息提取出
                            if r2 - r1 > 0:
                                cr.append((r1, r2, c1, c2))
                        # 这里注意需要把合并单元格的信息提取出再拆分
                        merge_cr = deepcopy(cr)
                        for r in cr:
                            stat_worksheet.unmerge_cells(start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])
                        stat_worksheet.delete_rows(3, stat_worksheet.max_row)
                        stat_worksheet.cell(2, 1, "")
                        stat_worksheet.cell(2, 2, "")
                    # data_only = True 保证公式正确计算出
                    workbook = openpyxl.load_workbook(path + "/" + file, data_only=True)
                    worksheet = workbook['每日进度']
                    for rows in list(worksheet.rows)[2:worksheet.max_row]:
                        if rows[11].value is not None and rows[11].value > 0:
                            for index in range(len(rows[0:worksheet.max_row])):
                                cell = rows[index]
                                if index == 0:
                                    temp_cell = worksheet.cell(2, 1)
                                    stat_cell = stat_worksheet.cell(stat_worksheet_init_row, index + 1)
                                    stat_cell.value = temp_cell.value
                                    stat_cell.border = copy(temp_cell.border)
                                    stat_cell.font = copy(temp_cell.font)
                                    stat_cell.fill = copy(temp_cell.fill)
                                    stat_cell.alignment = copy(temp_cell.alignment)
                                    stat_cell.protection = copy(temp_cell.protection)
                                    stat_cell.number_format = temp_cell.number_format
                                    print(temp_cell.value)
                                    # stat_worksheet.cell(stat_worksheet_init_row, index + 1, worksheet.cell(2, 1).value)
                                elif index == 1:
                                    temp_cell = worksheet.cell(2, 2)
                                    stat_cell = stat_worksheet.cell(stat_worksheet_init_row, index + 1)
                                    stat_cell.value = temp_cell.value
                                    stat_cell.border = copy(temp_cell.border)
                                    stat_cell.font = copy(temp_cell.font)
                                    stat_cell.fill = copy(temp_cell.fill)
                                    stat_cell.alignment = copy(temp_cell.alignment)
                                    stat_cell.protection = copy(temp_cell.protection)
                                    stat_cell.number_format = temp_cell.number_format
                                    # stat_worksheet.cell(stat_worksheet_init_row, index + 1, worksheet.cell(2, 2).value)
                                else:
                                    stat_cell = stat_worksheet.cell(stat_worksheet_init_row, index + 1)
                                    stat_cell.value = cell.value
                                    stat_cell.border = copy(cell.border)
                                    stat_cell.font = copy(cell.font)
                                    stat_cell.fill = copy(cell.fill)
                                    stat_cell.alignment = copy(cell.alignment)
                                    stat_cell.protection = copy(cell.protection)
                                    stat_cell.number_format = cell.number_format
                                    # stat_worksheet.cell(stat_worksheet_init_row, index + 1, cell.value)
                                    print(cell.value, end=" ")
                            stat_worksheet_init_row = stat_worksheet_init_row + 1
                            print()
        # Use a breakpoint in the code line below to debug your script.
        print(f'{name}')  # Press <no shortcut> to toggle the breakpoint.
        stat_workbook.save(filename=stat_file)
    except:
        print("程序执行出错了，可能未传入文件夹路径。")
        print('------ END ------')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print("------ START ------")
    print_hi('------ END ------')
